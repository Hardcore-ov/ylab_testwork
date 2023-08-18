import os
from datetime import datetime, timedelta
from uuid import UUID

from openpyxl import load_workbook

current_directory = os.path.dirname(__file__)
PATH_FILE_EXCEL = os.path.join(current_directory, '../../admin/Menu.xlsx')


def is_change(delay: int = 15) -> bool:

    statinfo = os.stat(PATH_FILE_EXCEL)
    last_modification_time = datetime.fromtimestamp(statinfo.st_mtime)
    delta_time = datetime.now() - last_modification_time
    return delta_time <= timedelta(seconds=delay)


def is_id(cell: str) -> bool:
    try:
        UUID(cell)
        return True
    except ValueError:
        return False


def get_data_from_excel_file():
    wb = load_workbook(PATH_FILE_EXCEL, data_only=True)
    ws = wb.active
    try:
        menus = dict()
        submenus = dict()
        dishes = dict()
        menu_id = None
        submenu_id = None
        for row in ws.iter_rows(values_only=True):
            if row[0] and is_id(row[0]):
                menus[str(UUID(row[0]))] = {
                    'title': row[1],
                    'description': row[2]
                }
                menu_id = str(UUID(row[0]))
                submenu_id = None
            if row[1] and is_id(row[1]):
                submenus[str(UUID(row[1]))] = {
                    'title': row[2],
                    'description': row[3],
                    'menu_id': menu_id
                }
                submenu_id = str(UUID(row[1]))
            if row[2] and is_id(row[2]):
                if row[6] is not None:
                    try:
                        int(row[6])
                    except ValueError:
                        return False
                    discount = int(row[6])
                else:
                    discount = 0
                dishes[str(UUID(row[2]))] = {
                    'title': row[3],
                    'description': row[4],
                    'price': row[5],
                    'discount': discount,
                    'submenu_id': submenu_id
                }
        return menus, submenus, dishes
    except Exception as ex:
        return ex
    finally:
        wb.close()
