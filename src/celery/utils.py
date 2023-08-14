import os

from openpyxl import load_workbook

current_directory = os.path.dirname(__file__)
PATH_FILE_EXCEL = os.path.join(current_directory, '../../admin/Menu.xlsx')

wb = load_workbook(PATH_FILE_EXCEL, data_only=True)
sheet = wb.active


def create_dict(row_iter):
    data = {}
    for row in row_iter:
        if row[0] is not None:
            data['title'] = row[1]
            data['description'] = row[2]
        else:
            continue
    return data


data_dict = create_dict(sheet.iter_rows(values_only=True))

print(data_dict)
